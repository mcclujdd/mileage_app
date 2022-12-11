# take data from .csv and input to appropriate cells in .xlsx template

from openpyxl import load_workbook
import csv
import shutil as sh
import decimal
from datetime import date
import _config as config

def mi2xlsx():
  year = date.today().year
  month = date.today().month
  prior_year = year - 1
  prior_month = month - 1
  src = "/private/var/mobile/Library/Mobile Documents/iCloud~com~omz-software~Pythonista3/Documents/mileage_2.0/mileage_expense_report.xlsx"

  name = config.name
  new_mileage_report = f'{name}_{prior_month}_{year}_mileage.xlsx'
  prior_year_final_report = f'{name}_12_{prior_year}_mileage.xlsx'
  path = config.cloud_path

  if month == 1:
    prior_month = 12
    sh.copy(src, prior_year_final_report)
    report = prior_year_final_report
    log_file = config.prior_output_file
  else:
    sh.copy(src, new_mileage_report)
    report = new_mileage_report
    log_file = config.output_file

  wb = load_workbook(filename = report)
  ws = wb['Mileage']

  row = 6
  with open(log_file, 'r') as file:
    csvreader = csv.reader(file)
    header = next(csvreader)
    tracked_sum = 0

    ws['J3'] = ws['F51'] = date.today().strftime('%m/%d/%y')[0:]
    for trip in csvreader:
      if trip[0][:2].strip() == str(prior_month):
        ws[f'B{row}'] = trip[0].strip()[0:]
        ws[f'D{row}'] = int(trip[3].strip())
        ws[f'F{row}'] = int(trip[4].strip())
        ws[f'H{row}'] = int(trip[5].strip())
        ws[f'J{row}'] = config.locations.get(trip[1].strip())+' to '+config.locations.get(trip[2].strip())
        row += 1
        tracked_sum += int(trip[5].strip())
    ws['H50'] = tracked_sum
    ws['J51'] = round(tracked_sum*decimal.Decimal(float(0.585)), 2)
    #only available in openpyxl v2.5^
    #ws.delete_rows(49, row+1)


  #save to cloud path
  wb.save(report)
  sh.move(report, path+report)
  wb.close()
  print(f'file saved successfully')

if __name__ == '__main__':
  mi2xlsx()
